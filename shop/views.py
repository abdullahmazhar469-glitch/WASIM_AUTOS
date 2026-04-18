from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.conf import settings
from .models import Product, Order, Cart, CartItem, Customer, Category, Review, OfflineSale
from .forms import RegisterForm, LoginForm, ReviewForm
import uuid
import json
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO

# ============================================
# CART FUNCTIONS
# ============================================

def get_cart(request):
    session_id = request.session.get('cart_session_id')
    if not session_id:
        session_id = str(uuid.uuid4())
        request.session['cart_session_id'] = session_id
        cart = Cart.objects.create(session_id=session_id)
    else:
        cart, created = Cart.objects.get_or_create(session_id=session_id)
    
    if request.user.is_authenticated and hasattr(request.user, 'customer'):
        customer_cart, created = Cart.objects.get_or_create(customer=request.user.customer)
        for item in cart.items.all():
            existing_item = customer_cart.items.filter(product=item.product).first()
            if existing_item:
                existing_item.quantity += item.quantity
                existing_item.save()
            else:
                item.cart = customer_cart
                item.save()
        cart.items.all().delete()
        cart = customer_cart
        request.session['cart_session_id'] = None
    
    return cart

def get_cart_count(request):
    cart = get_cart(request)
    return cart.items.aggregate(total=Sum('quantity'))['total'] or 0

# ============================================
# USER AUTHENTICATION
# ============================================

def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, '✅ Registration successful! Welcome to WASIM AUTOS!')
            return redirect('home')
        else:
            messages.error(request, '❌ Please correct the errors below.')
    else:
        form = RegisterForm()
    return render(request, 'register.html', {'form': form})

def user_login(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                messages.success(request, f'✅ Welcome back {username}!')
                return redirect('home')
            else:
                messages.error(request, '❌ Invalid username or password.')
    else:
        form = LoginForm()
    return render(request, 'login_user.html', {'form': form})

@login_required
def user_logout(request):
    logout(request)
    messages.success(request, '✅ Logged out successfully!')
    return redirect('home')

@login_required
def profile(request):
    customer = request.user.customer
    orders = Order.objects.filter(customer=customer).order_by('-order_date')
    cart_count = get_cart_count(request)
    return render(request, 'profile.html', {
        'customer': customer,
        'orders': orders,
        'cart_count': cart_count
    })

# ============================================
# HOME & PRODUCTS
# ============================================

def home(request):
    products = Product.objects.all()
    categories = Category.objects.all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(part_number__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(barcode__icontains=search_query)
        )
    
    # Category filter
    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)
    
    # Price filter
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    if min_price:
        products = products.filter(price__gte=min_price)
    if max_price:
        products = products.filter(price__lte=max_price)
    
    # Stock filter
    stock_filter = request.GET.get('stock')
    if stock_filter == 'in_stock':
        products = products.filter(stock__gt=0)
    elif stock_filter == 'low_stock':
        products = products.filter(stock__lte=5, stock__gt=0)
    elif stock_filter == 'out_of_stock':
        products = products.filter(stock=0)
    
    # Sort
    sort_by = request.GET.get('sort', '-id')
    products = products.order_by(sort_by)
    
    cart_count = get_cart_count(request)
    
    return render(request, 'home.html', {
        'products': products,
        'categories': categories,
        'search_query': search_query,
        'cart_count': cart_count,
        'min_price': min_price,
        'max_price': max_price,
        'sort_by': sort_by,
        'stock_filter': stock_filter,
    })

def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    related_products = Product.objects.filter(category=product.category).exclude(id=product_id)[:4]
    cart_count = get_cart_count(request)
    reviews = product.reviews.all()
    
    if request.method == 'POST' and request.user.is_authenticated:
        form = ReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.product = product
            review.customer = request.user.customer
            review.save()
            messages.success(request, '✅ Review added successfully!')
            return redirect('product_detail', product_id=product_id)
    else:
        form = ReviewForm()
    
    return render(request, 'product_detail.html', {
        'product': product,
        'related_products': related_products,
        'cart_count': cart_count,
        'reviews': reviews,
        'form': form
    })

def search_by_barcode(request):
    barcode = request.GET.get('barcode', '')
    if barcode:
        product = Product.objects.filter(barcode=barcode).first()
        if product:
            return JsonResponse({'success': True, 'product_id': product.id, 'name': product.name, 'price': str(product.price)})
    return JsonResponse({'success': False})

def add_review(request, product_id):
    if request.method == 'POST' and request.user.is_authenticated:
        product = get_object_or_404(Product, id=product_id)
        rating = int(request.POST.get('rating', 0))
        comment = request.POST.get('comment', '')
        
        Review.objects.create(
            product=product,
            customer=request.user.customer,
            rating=rating,
            comment=comment
        )
        messages.success(request, '✅ Review added successfully!')
    return redirect('product_detail', product_id=product_id)

# ============================================
# CART & CHECKOUT
# ============================================

def add_to_cart(request, product_id):
    if request.method == 'POST':
        cart = get_cart(request)
        product = get_object_or_404(Product, id=product_id)
        quantity = int(request.POST.get('quantity', 1))
        
        cart_item, created = CartItem.objects.get_or_create(
            cart=cart,
            product=product,
            defaults={'quantity': quantity}
        )
        if not created:
            cart_item.quantity += quantity
            cart_item.save()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            cart_count = get_cart_count(request)
            return JsonResponse({
                'success': True,
                'cart_count': cart_count,
                'message': f'✅ {product.name} added to cart!'
            })
        
        return redirect('cart')
    return redirect('home')

def cart_view(request):
    cart = get_cart(request)
    cart_items = cart.items.all()
    cart_count = get_cart_count(request)
    total_amount = cart.total_amount()
    
    return render(request, 'cart.html', {
        'cart_items': cart_items,
        'cart_count': cart_count,
        'total_amount': total_amount
    })

def update_cart(request, item_id):
    if request.method == 'POST':
        cart_item = get_object_or_404(CartItem, id=item_id)
        quantity = int(request.POST.get('quantity', 0))
        if quantity <= 0:
            cart_item.delete()
        else:
            cart_item.quantity = quantity
            cart_item.save()
        return redirect('cart')

def remove_from_cart(request, item_id):
    if request.method == 'POST':
        cart_item = get_object_or_404(CartItem, id=item_id)
        cart_item.delete()
        return redirect('cart')

def checkout(request):
    cart = get_cart(request)
    cart_items = cart.items.all()
    total_amount = cart.total_amount()
    cart_count = get_cart_count(request)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        payment_method = request.POST.get('payment_method', 'COD')
        
        customer = None
        if request.user.is_authenticated:
            customer = request.user.customer
        
        for item in cart_items:
            order = Order.objects.create(
                product=item.product,
                customer=customer,
                customer_name=name,
                email=email,
                phone=phone,
                address=address,
                quantity=item.quantity,
                total_price=item.total_price(),
                status='Pending',
                payment_method=payment_method
            )
            item.product.stock -= item.quantity
            item.product.save()
            
            # Send email notification
            if email:
                try:
                    send_mail(
                        subject=f'Order Confirmation - Order #{order.id}',
                        message=f'Dear {name},\n\nYour order has been placed successfully!\n\nOrder Details:\nOrder ID: {order.id}\nTotal: Rs. {item.total_price()}\nPayment Method: {payment_method}\n\nThank you for shopping with WASIM AUTOS!',
                        from_email=settings.EMAIL_HOST_USER,
                        recipient_list=[email],
                        fail_silently=True,
                    )
                except:
                    pass
        
        cart.items.all().delete()
        messages.success(request, '✅ Order placed successfully! We will contact you soon.')
        return redirect('home')
    
    return render(request, 'checkout.html', {
        'cart_items': cart_items,
        'total_amount': total_amount,
        'cart_count': cart_count
    })

def order_form(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cart_count = get_cart_count(request)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        qty = int(request.POST.get('quantity', 1))
        payment_method = request.POST.get('payment_method', 'COD')
        
        if qty > product.stock:
            return render(request, 'order.html', {
                'product': product,
                'error': '❌ Out of stock!',
                'cart_count': cart_count
            })
        
        order = Order.objects.create(
            customer_name=name,
            phone=phone,
            address=address,
            product=product,
            quantity=qty,
            total_price=product.price * qty,
            status='Pending',
            payment_method=payment_method
        )
        product.stock -= qty
        product.save()
        
        messages.success(request, '✅ Order placed successfully!')
        return render(request, 'invoice.html', {
            'order': order,
            'cart_count': cart_count
        })
    
    return render(request, 'order.html', {
        'product': product,
        'cart_count': cart_count
    })

# ============================================
# EXPORT REPORTS
# ============================================

def export_orders_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="orders.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    headers = ['ID', 'Customer', 'Email', 'Phone', 'Product', 'Quantity', 'Total', 'Status', 'Payment', 'Date']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    orders = Order.objects.all()
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.id)
        ws.cell(row=row, column=2, value=order.customer_name)
        ws.cell(row=row, column=3, value=order.email or '')
        ws.cell(row=row, column=4, value=order.phone)
        ws.cell(row=row, column=5, value=order.product.name)
        ws.cell(row=row, column=6, value=order.quantity)
        ws.cell(row=row, column=7, value=float(order.total_price))
        ws.cell(row=row, column=8, value=order.status)
        ws.cell(row=row, column=9, value=order.payment_method)
        ws.cell(row=row, column=10, value=str(order.order_date))
    
    wb.save(response)
    return response

def export_products_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    
    headers = ['ID', 'Name', 'Part Number', 'Category', 'Cost Price', 'Sale Price', 'Stock', 'Profit/Unit', 'Margin%']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    products = Product.objects.all()
    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.id)
        ws.cell(row=row, column=2, value=product.name)
        ws.cell(row=row, column=3, value=product.part_number or '')
        ws.cell(row=row, column=4, value=product.category.name if product.category else '')
        ws.cell(row=row, column=5, value=float(product.cost_price or 0))
        ws.cell(row=row, column=6, value=float(product.price))
        ws.cell(row=row, column=7, value=product.stock)
        ws.cell(row=row, column=8, value=float(product.profit_per_unit()))
        ws.cell(row=row, column=9, value=float(product.profit_margin()))
    
    wb.save(response)
    return response

import requests  # Add this at the top for SMS API
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from io import BytesIO
import qrcode
from datetime import datetime

# Add these imports at the top if not already there
from django.core.mail import EmailMultiAlternatives
from django.conf import settings

def checkout(request):
    cart = get_cart(request)
    cart_items = cart.items.all()
    total_amount = cart.total_amount()
    cart_count = get_cart_count(request)
    
    # Get or create customer for logged-in user
    customer = None
    if request.user.is_authenticated:
        try:
            customer = request.user.customer
        except:
            customer = Customer.objects.create(
                user=request.user,
                name=request.user.username,
                email=request.user.email or '',
                phone='',
                address=''
            )
    
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        payment_method = request.POST.get('payment_method', 'COD')
        
        # Update customer info if logged in
        if customer and request.user.is_authenticated:
            customer.name = name
            customer.email = email
            customer.phone = phone
            customer.address = address
            customer.save()
        
        orders_created = []
        
        for item in cart_items:
            order = Order.objects.create(
                product=item.product,
                customer=customer,
                customer_name=name,
                email=email,
                phone=phone,
                address=address,
                quantity=item.quantity,
                total_price=item.total_price(),
                status='Pending',
                payment_method=payment_method
            )
            item.product.stock -= item.quantity
            item.product.save()
            orders_created.append(order)
            
            # Send Email Invoice
            send_order_email(order, email, name)
            
            # Send SMS
            send_order_sms(phone, order)
        
        # Generate PDF Invoice
        pdf_buffer = generate_pdf_invoice(orders_created[0] if orders_created else None, name, email, phone, address, total_amount)
        
        cart.items.all().delete()
        messages.success(request, '✅ Order placed successfully! Check your email for invoice.')
        
        # Return PDF invoice download
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{orders_created[0].id}.pdf"'
        return response
    
    # Pre-fill form with customer data if available
    initial_data = {}
    if customer:
        initial_data = {
            'name': customer.name,
            'email': customer.email,
            'phone': customer.phone,
            'address': customer.address,
        }
    
    return render(request, 'checkout.html', {
        'cart_items': cart_items,
        'total_amount': total_amount,
        'cart_count': cart_count,
        'customer': customer,
        'initial': initial_data
    })


def send_order_email(order, to_email, customer_name):
    """Send HTML email invoice to customer"""
    try:
        # HTML email template
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .invoice {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: #f0c040; padding: 20px; text-align: center; }}
                .order-details {{ margin: 20px 0; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th, td {{ padding: 10px; border-bottom: 1px solid #ddd; text-align: left; }}
                .total {{ font-size: 18px; font-weight: bold; color: #f0c040; }}
            </style>
        </head>
        <body>
            <div class="invoice">
                <div class="header">
                    <h2>WASIM AUTOS</h2>
                    <p>Rahim Yar Khan, Pakistan</p>
                </div>
                <div class="order-details">
                    <h3>Order Confirmation #{order.id}</h3>
                    <p><strong>Dear {customer_name},</strong></p>
                    <p>Thank you for your order! Your order has been confirmed and will be processed soon.</p>
                    
                    <h4>Order Details:</h4>
                    <table>
                        <tr>
                            <th>Product</th>
                            <th>Quantity</th>
                            <th>Price</th>
                            <th>Total</th>
                        </tr>
                        <tr>
                            <td>{order.product.name}</td>
                            <td>{order.quantity}</td>
                            <td>Rs. {order.product.price}</td>
                            <td>Rs. {order.total_price}</td>
                        </tr>
                    </table>
                    
                    <p><strong>Payment Method:</strong> {order.payment_method}</p>
                    <p><strong>Shipping Address:</strong> {order.address}</p>
                    <p><strong>Phone:</strong> {order.phone}</p>
                    
                    <div class="total">
                        <p>Total Amount: Rs. {order.total_price}</p>
                    </div>
                    
                    <p>We will notify you once your order is shipped.</p>
                    <p>For any queries, contact us at +92 328 3042403</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        plain_text = strip_tags(html_content)
        
        email_msg = EmailMultiAlternatives(
            subject=f'Order Confirmation - WASIM AUTOS #{order.id}',
            body=plain_text,
            from_email=settings.EMAIL_HOST_USER,
            to=[to_email]
        )
        email_msg.attach_alternative(html_content, "text/html")
        email_msg.send()
        
        print(f"✅ Email sent to {to_email}")
        
    except Exception as e:
        print(f"❌ Email error: {str(e)}")


def send_order_sms(phone_number, order):
    """Send SMS to customer using Fast2SMS or any SMS API"""
    try:
        # Using Fast2SMS API (free for testing)
        # Sign up at https://www.fast2sms.com/ for API key
        
        API_KEY = "YOUR_FAST2SMS_API_KEY"  # Replace with your API key
        
        message = f"WASIM AUTOS: Order #{order.id} confirmed! Amount: Rs.{order.total_price}. We'll deliver soon. Thank you!"
        
        url = "https://www.fast2sms.com/dev/bulkV2"
        
        payload = {
            "sender_id": "TXTIND",
            "message": message,
            "language": "english",
            "route": "v3",
            "numbers": phone_number,
        }
        
        headers = {
            'authorization': API_KEY,
            'Content-Type': "application/json"
        }
        
        # Uncomment to send actual SMS
        # response = requests.post(url, json=payload, headers=headers)
        # print(f"SMS sent: {response.json()}")
        
        print(f"✅ SMS would be sent to {phone_number}: {message}")
        
    except Exception as e:
        print(f"❌ SMS error: {str(e)}")


def generate_pdf_invoice(order, customer_name, email, phone, address, total_amount):
    """Generate PDF invoice for download"""
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    p.setFont("Helvetica-Bold", 20)
    p.setFillColorRGB(0.94, 0.75, 0.25)  # #f0c040 color
    p.drawString(50, height - 50, "WASIM AUTOS")
    
    p.setFont("Helvetica", 10)
    p.setFillColorRGB(0.5, 0.5, 0.5)
    p.drawString(50, height - 70, "Rahim Yar Khan, Pakistan")
    p.drawString(50, height - 85, "Phone: +92 328 3042403")
    
    # Invoice Title
    p.setFont("Helvetica-Bold", 16)
    p.setFillColorRGB(0, 0, 0)
    p.drawString(50, height - 120, f"TAX INVOICE #{order.id if order else 'N/A'}")
    
    # Order Details
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, height - 150, "Order Details:")
    
    p.setFont("Helvetica", 10)
    p.drawString(50, height - 170, f"Date: {datetime.now().strftime('%d %b %Y, %I:%M %p')}")
    p.drawString(50, height - 185, f"Customer: {customer_name}")
    p.drawString(50, height - 200, f"Phone: {phone}")
    p.drawString(50, height - 215, f"Email: {email}")
    
    # Address
    p.drawString(50, height - 240, f"Shipping Address: {address}")
    
    # Product Table Header
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, height - 280, "Product")
    p.drawString(300, height - 280, "Qty")
    p.drawString(350, height - 280, "Price")
    p.drawString(450, height - 280, "Total")
    
    # Product Row
    if order:
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 300, order.product.name[:40])
        p.drawString(300, height - 300, str(order.quantity))
        p.drawString(350, height - 300, f"Rs. {order.product.price}")
        p.drawString(450, height - 300, f"Rs. {order.total_price}")
    
    # Total
    p.setFont("Helvetica-Bold", 12)
    p.setFillColorRGB(0.94, 0.75, 0.25)
    p.drawString(350, height - 350, f"Total Amount: Rs. {total_amount}")
    
    # Footer
    p.setFont("Helvetica", 8)
    p.setFillColorRGB(0.5, 0.5, 0.5)
    p.drawString(50, 50, "Thank you for shopping with WASIM AUTOS!")
    p.drawString(50, 35, "For support: info@wasimautos.com | +92 328 3042403")
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    return buffer


def order_form(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cart_count = get_cart_count(request)
    
    # Get or create customer for logged-in user
    customer = None
    if request.user.is_authenticated:
        try:
            customer = request.user.customer
        except:
            customer = Customer.objects.create(
                user=request.user,
                name=request.user.username,
                email=request.user.email or '',
                phone='',
                address=''
            )
    
    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        qty = int(request.POST.get('quantity', 1))
        payment_method = request.POST.get('payment_method', 'COD')
        email = request.POST.get('email', '')
        
        if qty > product.stock:
            messages.error(request, '❌ Out of stock!')
            return render(request, 'order.html', {
                'product': product,
                'error': '❌ Out of stock!',
                'cart_count': cart_count,
                'customer': customer
            })
        
        # Update customer info if logged in
        if customer and request.user.is_authenticated:
            customer.name = name
            customer.phone = phone
            customer.address = address
            customer.email = email
            customer.save()
        
        order = Order.objects.create(
            customer_name=name,
            phone=phone,
            address=address,
            product=product,
            customer=customer,
            email=email,
            quantity=qty,
            total_price=product.price * qty,
            status='Pending',
            payment_method=payment_method
        )
        product.stock -= qty
        product.save()
        
        # Send Email
        send_order_email(order, email, name)
        
        # Send SMS
        send_order_sms(phone, order)
        
        messages.success(request, f'✅ Order #{order.id} placed successfully! Invoice sent to your email.')
        
        # Generate and return PDF invoice
        pdf_buffer = generate_pdf_invoice(order, name, email, phone, address, order.total_price)
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{order.id}.pdf"'
        return response
    
    return render(request, 'order.html', {
        'product': product,
        'cart_count': cart_count,
        'customer': customer
    })